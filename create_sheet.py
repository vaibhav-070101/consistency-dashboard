#!/usr/bin/env python3
"""Generate the Consistency Dashboard Excel spreadsheet.

Creates a beautiful, formula-driven .xlsx file optimized for
Google Sheets on iPhone. Each habit has a weekly target, and
consistency % is calculated relative to that target.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import date
import calendar
import math

# ── Config ────────────────────────────────────────────────────────────
YEAR = 2026
OUTPUT = "Consistency_Dashboard_2026.xlsx"

HABITS = [
    {"name": "Gym",        "target": 5},
    {"name": "Speaking",   "target": 5},
    {"name": "Wake Early", "target": 7},
    {"name": "Eat Clean",  "target": 7},
    {"name": "Calories",   "target": 7},
    {"name": "Coding",     "target": 5},
]

NH = len(HABITS)
MONTH_NAMES = list(calendar.month_name)[1:]  # January..December

# Column mapping: A=label, B=extra, C..H=habits, I=average
HC = 3  # first habit column (C)
AVG_C = HC + NH  # column I

# Row layout within each month sheet
R_TITLE = 1
R_CONFIG_LABEL = 3
R_WEEKLY_TITLE = 5
R_WEEKLY_HDR = 6
R_W1 = 7  # Week 1 data row; weeks go 7..11
R_MONTHLY_TITLE = 13
R_MONTHLY_DATA = 14
R_LOG_TITLE = 16
R_LOG_HDR = 17
R_LOG_START = 18  # first day-of-month row

# ── Colors ────────────────────────────────────────────────────────────
C_DARK_GREEN  = "1B5E20"
C_MED_GREEN   = "2E7D32"
C_GREEN       = "4CAF50"
C_LIGHT_GREEN = "C8E6C9"
C_VLIGHT_GREEN = "E8F5E9"
C_WHITE       = "FFFFFF"
C_LGRAY       = "F5F5F5"
C_GRAY        = "BDBDBD"
C_DGRAY       = "616161"
C_TEXT        = "212121"
C_RED         = "E53935"
C_LRED        = "FFCDD2"
C_ORANGE      = "FF9800"
C_LORANGE     = "FFF3E0"

# ── Reusable Styles ──────────────────────────────────────────────────
F_TITLE    = Font(name="Calibri", size=18, bold=True, color=C_DARK_GREEN)
F_SECTION  = Font(name="Calibri", size=12, bold=True, color=C_WHITE)
F_HDR      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
F_SUBHDR   = Font(name="Calibri", size=10, bold=True, color=C_DARK_GREEN)
F_DATA     = Font(name="Calibri", size=11, color=C_TEXT)
F_DATA_DIM = Font(name="Calibri", size=10, color=C_DGRAY)
F_PCT      = Font(name="Calibri", size=11, bold=True, color=C_MED_GREEN)
F_PCT_BIG  = Font(name="Calibri", size=14, bold=True, color=C_MED_GREEN)
F_DASH_TITLE = Font(name="Calibri", size=22, bold=True, color=C_DARK_GREEN)
F_DASH_SUB = Font(name="Calibri", size=11, color=C_DGRAY)
F_DASH_HDR = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
F_DASH_METRIC = Font(name="Calibri", size=28, bold=True, color=C_MED_GREEN)
F_DASH_LABEL = Font(name="Calibri", size=10, color=C_DGRAY)

FILL_DARK_GREEN = PatternFill("solid", fgColor=C_MED_GREEN)
FILL_GREEN      = PatternFill("solid", fgColor=C_GREEN)
FILL_LGREEN     = PatternFill("solid", fgColor=C_LIGHT_GREEN)
FILL_VLGREEN    = PatternFill("solid", fgColor=C_VLIGHT_GREEN)
FILL_WHITE      = PatternFill("solid", fgColor=C_WHITE)
FILL_LGRAY      = PatternFill("solid", fgColor=C_LGRAY)
FILL_LRED       = PatternFill("solid", fgColor=C_LRED)
FILL_LORANGE    = PatternFill("solid", fgColor=C_LORANGE)

THIN = Side(style="thin", color=C_GRAY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left", vertical="center")
PCT_FMT = "0%"

# ── Helper functions ─────────────────────────────────────────────────

def cell(ws, row, col, value=None, font=None, fill=None, align=CENTER,
         border=BORDER, fmt=None, merge_end_col=None):
    """Write a styled cell."""
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    if fmt:    c.number_format = fmt
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    return c


def style_range(ws, row, col_start, col_end, **kwargs):
    """Apply style to a range of cells in a row."""
    for c in range(col_start, col_end + 1):
        cell(ws, row, c, **kwargs)


def week_chunks(num_days):
    """Split month days into 7-day week chunks. Returns list of (start_day, end_day) 1-indexed."""
    chunks = []
    d = 1
    while d <= num_days:
        end = min(d + 6, num_days)
        chunks.append((d, end))
        d = end + 1
    return chunks


def adjusted_target(full_target, days_in_chunk):
    """Scale weekly target for partial weeks."""
    if days_in_chunk >= 7:
        return full_target
    return max(1, round(full_target * days_in_chunk / 7))


# ── Build workbook ───────────────────────────────────────────────────

wb = Workbook()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CREATE MONTH SHEETS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

for mi in range(12):
    month_num = mi + 1
    mname = MONTH_NAMES[mi]
    num_days = calendar.monthrange(YEAR, month_num)[1]
    weeks = week_chunks(num_days)
    num_weeks = len(weeks)

    if mi == 0:
        ws = wb.active
        ws.title = mname
    else:
        ws = wb.create_sheet(mname)

    ws.sheet_properties.tabColor = C_MED_GREEN

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    for h in range(NH):
        ws.column_dimensions[get_column_letter(HC + h)].width = 13
    ws.column_dimensions[get_column_letter(AVG_C)].width = 13

    # ── Row 1: Title ─────────────────────────────────────────────
    cell(ws, R_TITLE, 1, f"{mname.upper()} {YEAR}", font=F_TITLE,
         fill=None, align=LEFT, border=None, merge_end_col=AVG_C)

    # ── Row 2: Subtitle ─────────────────────────────────────────
    cell(ws, 2, 1, "CONSISTENCY TRACKER", font=F_DASH_SUB,
         fill=None, align=LEFT, border=None, merge_end_col=AVG_C)

    # ── Row 3: Config/Targets label ─────────────────────────────
    cell(ws, R_CONFIG_LABEL, 1, "Targets →", font=F_SUBHDR,
         fill=FILL_VLGREEN, align=LEFT)
    cell(ws, R_CONFIG_LABEL, 2, "", fill=FILL_VLGREEN)
    for h_i, h in enumerate(HABITS):
        col = HC + h_i
        cell(ws, R_CONFIG_LABEL, col, f"{h['target']}/wk",
             font=F_DATA_DIM, fill=FILL_VLGREEN)
    cell(ws, R_CONFIG_LABEL, AVG_C, "", fill=FILL_VLGREEN)

    # ── Row 5: Weekly Summary title ──────────────────────────────
    cell(ws, R_WEEKLY_TITLE, 1, "WEEKLY BREAKDOWN", font=F_SECTION,
         fill=FILL_DARK_GREEN, align=LEFT, merge_end_col=AVG_C)
    for c in range(2, AVG_C + 1):
        cell(ws, R_WEEKLY_TITLE, c, fill=FILL_DARK_GREEN, font=F_SECTION)

    # ── Row 6: Weekly column headers ─────────────────────────────
    cell(ws, R_WEEKLY_HDR, 1, "Week", font=F_HDR, fill=FILL_GREEN)
    cell(ws, R_WEEKLY_HDR, 2, "Days", font=F_HDR, fill=FILL_GREEN)
    for h_i, h in enumerate(HABITS):
        col = HC + h_i
        cell(ws, R_WEEKLY_HDR, col, h["name"], font=F_HDR, fill=FILL_GREEN)
    cell(ws, R_WEEKLY_HDR, AVG_C, "Avg", font=F_HDR, fill=FILL_GREEN)

    # ── Rows 7-11: Week data ────────────────────────────────────
    for wi, (d_start, d_end) in enumerate(weeks):
        row = R_W1 + wi
        days_in_wk = d_end - d_start + 1
        log_row_start = R_LOG_START + d_start - 1
        log_row_end = R_LOG_START + d_end - 1

        cell(ws, row, 1, f"Week {wi+1}", font=F_DATA, fill=FILL_LGRAY, align=LEFT)
        cell(ws, row, 2, days_in_wk, font=F_DATA_DIM, fill=FILL_LGRAY)

        for h_i, h in enumerate(HABITS):
            col = HC + h_i
            col_letter = get_column_letter(col)
            adj_t = adjusted_target(h["target"], days_in_wk)
            formula = (
                f'=MIN(COUNTIF({col_letter}{log_row_start}:{col_letter}{log_row_end},1)'
                f'/{adj_t},1)'
            )
            cell(ws, row, col, font=F_PCT, fill=FILL_LGRAY, fmt=PCT_FMT)
            ws.cell(row=row, column=col).value = formula

        # Week average = average of all habit percentages this week
        habit_cells = ",".join(
            f"{get_column_letter(HC + i)}{row}" for i in range(NH)
        )
        cell(ws, row, AVG_C, f"=AVERAGE({habit_cells})", font=F_PCT,
             fill=FILL_VLGREEN, fmt=PCT_FMT)

    # Blank out unused week rows (if fewer than 5 weeks)
    for wi in range(num_weeks, 5):
        row = R_W1 + wi
        for c in range(1, AVG_C + 1):
            cell(ws, row, c, "", font=F_DATA, fill=FILL_WHITE)

    # ── Row 13: Monthly title ────────────────────────────────────
    cell(ws, R_MONTHLY_TITLE, 1, "MONTHLY CONSISTENCY", font=F_SECTION,
         fill=FILL_DARK_GREEN, align=LEFT, merge_end_col=AVG_C)
    for c in range(2, AVG_C + 1):
        cell(ws, R_MONTHLY_TITLE, c, fill=FILL_DARK_GREEN, font=F_SECTION)

    # ── Row 14: Monthly averages ─────────────────────────────────
    cell(ws, R_MONTHLY_DATA, 1, "Average", font=F_SUBHDR,
         fill=FILL_VLGREEN, align=LEFT)
    cell(ws, R_MONTHLY_DATA, 2, "", fill=FILL_VLGREEN)

    week_rows_range = f"{R_W1}:{R_W1 + num_weeks - 1}"
    for h_i in range(NH):
        col = HC + h_i
        cl = get_column_letter(col)
        first_wk = R_W1
        last_wk = R_W1 + num_weeks - 1
        formula = f"=AVERAGE({cl}{first_wk}:{cl}{last_wk})"
        cell(ws, R_MONTHLY_DATA, col, formula, font=F_PCT_BIG,
             fill=FILL_VLGREEN, fmt=PCT_FMT)

    # Overall monthly average
    habit_cells = ",".join(
        f"{get_column_letter(HC + i)}{R_MONTHLY_DATA}" for i in range(NH)
    )
    cell(ws, R_MONTHLY_DATA, AVG_C,
         f"=AVERAGE({habit_cells})", font=F_PCT_BIG,
         fill=FILL_LGREEN, fmt=PCT_FMT)

    # ── Row 16: Daily Log title ──────────────────────────────────
    cell(ws, R_LOG_TITLE, 1, "DAILY LOG  ·  Enter 1 when done", font=F_SECTION,
         fill=FILL_DARK_GREEN, align=LEFT, merge_end_col=AVG_C)
    for c in range(2, AVG_C + 1):
        cell(ws, R_LOG_TITLE, c, fill=FILL_DARK_GREEN, font=F_SECTION)

    # ── Row 17: Daily log column headers ─────────────────────────
    cell(ws, R_LOG_HDR, 1, "Date", font=F_HDR, fill=FILL_GREEN)
    cell(ws, R_LOG_HDR, 2, "Day", font=F_HDR, fill=FILL_GREEN)
    for h_i, h in enumerate(HABITS):
        cell(ws, R_LOG_HDR, HC + h_i, h["name"], font=F_HDR, fill=FILL_GREEN)
    cell(ws, R_LOG_HDR, AVG_C, "Score", font=F_HDR, fill=FILL_GREEN)

    # ── Rows 18+: Daily data rows ───────────────────────────────
    for d in range(1, num_days + 1):
        row = R_LOG_START + d - 1
        dt = date(YEAR, month_num, d)
        day_name = dt.strftime("%a")

        is_weekend = dt.weekday() >= 5
        bg = FILL_LGRAY if is_weekend else FILL_WHITE

        cell(ws, row, 1, dt.strftime("%b %d"), font=F_DATA, fill=bg, align=LEFT)
        cell(ws, row, 2, day_name, font=F_DATA_DIM, fill=bg)

        for h_i in range(NH):
            col = HC + h_i
            cell(ws, row, col, None, font=F_DATA, fill=bg)

        # Daily score: count of habits done / total habits
        habit_range = (
            f"{get_column_letter(HC)}{row}:{get_column_letter(HC + NH - 1)}{row}"
        )
        cell(ws, row, AVG_C,
             f'=IF(COUNTIF({habit_range},">=1")>0,'
             f'COUNTIF({habit_range},">=1")/{NH},"")',
             font=F_PCT, fill=bg, fmt=PCT_FMT)

    # ── Week separator highlighting ──────────────────────────────
    for wi, (d_start, d_end) in enumerate(weeks):
        if wi == 0:
            continue
        sep_row = R_LOG_START + d_start - 1
        for c in range(1, AVG_C + 1):
            existing = ws.cell(row=sep_row, column=c)
            existing.border = Border(
                left=THIN, right=THIN, bottom=THIN,
                top=Side(style="medium", color=C_MED_GREEN)
            )

    # ── Conditional formatting: green fill when habit cell = 1 ───
    for h_i in range(NH):
        col_letter = get_column_letter(HC + h_i)
        data_range = f"{col_letter}{R_LOG_START}:{col_letter}{R_LOG_START + num_days - 1}"
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="equal", formula=["1"],
                       fill=PatternFill("solid", fgColor=C_LIGHT_GREEN),
                       font=Font(bold=True, color=C_DARK_GREEN))
        )

    # Conditional formatting for weekly percentages
    for wi in range(num_weeks):
        row = R_W1 + wi
        pct_range = f"{get_column_letter(HC)}{row}:{get_column_letter(AVG_C)}{row}"
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
                       fill=PatternFill("solid", fgColor=C_LIGHT_GREEN))
        )
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="between", formula=["0.5", "0.7999"],
                       fill=PatternFill("solid", fgColor=C_LORANGE))
        )
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="lessThan", formula=["0.5"],
                       fill=PatternFill("solid", fgColor=C_LRED))
        )

    # Monthly consistency conditional formatting
    monthly_range = f"{get_column_letter(HC)}{R_MONTHLY_DATA}:{get_column_letter(AVG_C)}{R_MONTHLY_DATA}"
    ws.conditional_formatting.add(
        monthly_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
                   fill=PatternFill("solid", fgColor=C_LIGHT_GREEN))
    )

    # ── Freeze panes: freeze header row + date column ────────────
    ws.freeze_panes = ws.cell(row=R_LOG_START, column=HC)

    # ── Print setup ──────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr = None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CREATE DASHBOARD SHEET
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

dash = wb.create_sheet("Dashboard", 0)
dash.sheet_properties.tabColor = C_DARK_GREEN

dash.column_dimensions["A"].width = 18
dash.column_dimensions["B"].width = 10
for i in range(12):
    dash.column_dimensions[get_column_letter(3 + i)].width = 10
dash.column_dimensions[get_column_letter(15)].width = 12

# Title
cell(dash, 1, 1, "THE CONSISTENCY DASHBOARD", font=F_DASH_TITLE,
     fill=None, align=LEFT, border=None, merge_end_col=15)
cell(dash, 2, 1, f"{YEAR}  ·  Your year of building unbreakable habits",
     font=F_DASH_SUB, fill=None, align=LEFT, border=None, merge_end_col=15)

# ── Stat cards row (row 4) ───────────────────────────────────────
stat_labels = ["Overall Consistency", "Best Month", "Habits Tracked", "Year Progress"]
stat_cols = [1, 4, 7, 10]

for idx, (label, sc) in enumerate(zip(stat_labels, stat_cols)):
    for c in range(sc, sc + 3):
        cell(dash, 4, c, fill=FILL_VLGREEN)
        cell(dash, 5, c, fill=FILL_VLGREEN)
    cell(dash, 5, sc, label, font=F_DASH_LABEL, fill=FILL_VLGREEN,
         align=CENTER, border=None, merge_end_col=sc + 2)

# Overall consistency formula (average of all month averages that have data)
all_month_refs = []
for mi, mname in enumerate(MONTH_NAMES):
    all_month_refs.append(f"'{mname}'!{get_column_letter(AVG_C)}{R_MONTHLY_DATA}")
avg_formula = "=IFERROR(AVERAGE(" + ",".join(all_month_refs) + '),"")'
cell(dash, 4, 1, avg_formula, font=F_DASH_METRIC, fill=FILL_VLGREEN,
     fmt=PCT_FMT, border=None, merge_end_col=3)

# Best month: we'll just show "—" as a placeholder (complex formula)
cell(dash, 4, 4, "—", font=F_DASH_METRIC, fill=FILL_VLGREEN,
     border=None, merge_end_col=6)

cell(dash, 4, 7, NH, font=F_DASH_METRIC, fill=FILL_VLGREEN,
     border=None, merge_end_col=9)

# Year progress (what % of year has passed)
cell(dash, 4, 10, f'=ROUND((TODAY()-DATE({YEAR},1,1))/365,2)',
     font=F_DASH_METRIC, fill=FILL_VLGREEN, fmt=PCT_FMT,
     border=None, merge_end_col=12)

# ── Monthly breakdown table (row 7+) ────────────────────────────
cell(dash, 7, 1, "MONTHLY BREAKDOWN", font=F_SECTION,
     fill=FILL_DARK_GREEN, align=LEFT, merge_end_col=15)
for c in range(2, 16):
    cell(dash, 7, c, fill=FILL_DARK_GREEN, font=F_SECTION)

# Headers
cell(dash, 8, 1, "Habit", font=F_HDR, fill=FILL_GREEN, align=LEFT)
cell(dash, 8, 2, "Target", font=F_HDR, fill=FILL_GREEN)
for mi in range(12):
    cell(dash, 8, 3 + mi, MONTH_NAMES[mi][:3], font=F_HDR, fill=FILL_GREEN)
cell(dash, 8, 15, "Year Avg", font=F_HDR, fill=FILL_GREEN)

# Per-habit rows
for h_i, h in enumerate(HABITS):
    row = 9 + h_i
    cell(dash, row, 1, h["name"], font=F_DATA, fill=FILL_WHITE, align=LEFT)
    cell(dash, row, 2, f"{h['target']}/wk", font=F_DATA_DIM, fill=FILL_LGRAY)

    month_cells = []
    for mi in range(12):
        mname = MONTH_NAMES[mi]
        col = 3 + mi
        habit_col_letter = get_column_letter(HC + h_i)
        ref = f"=IFERROR('{mname}'!{habit_col_letter}{R_MONTHLY_DATA},\"\")"
        cell(dash, row, col, ref, font=F_PCT, fill=FILL_WHITE, fmt=PCT_FMT)
        month_cells.append(f"{get_column_letter(col)}{row}")

    # Year average
    year_range = f"{get_column_letter(3)}{row}:{get_column_letter(14)}{row}"
    cell(dash, row, 15, f"=IFERROR(AVERAGE({year_range}),\"\")",
         font=F_PCT_BIG, fill=FILL_VLGREEN, fmt=PCT_FMT)

# Overall row
overall_row = 9 + NH
cell(dash, overall_row, 1, "OVERALL", font=Font(name="Calibri", size=11, bold=True, color=C_WHITE),
     fill=FILL_DARK_GREEN, align=LEFT)
cell(dash, overall_row, 2, "", fill=FILL_DARK_GREEN)
for mi in range(12):
    col = 3 + mi
    col_range = f"{get_column_letter(col)}{9}:{get_column_letter(col)}{9 + NH - 1}"
    cell(dash, overall_row, col, f"=IFERROR(AVERAGE({col_range}),\"\")",
         font=Font(name="Calibri", size=11, bold=True, color=C_WHITE),
         fill=FILL_DARK_GREEN, fmt=PCT_FMT)

year_range = f"{get_column_letter(3)}{overall_row}:{get_column_letter(14)}{overall_row}"
cell(dash, overall_row, 15, f"=IFERROR(AVERAGE({year_range}),\"\")",
     font=Font(name="Calibri", size=14, bold=True, color=C_WHITE),
     fill=FILL_DARK_GREEN, fmt=PCT_FMT)

# Conditional formatting for the monthly breakdown
data_range = f"C9:{get_column_letter(15)}{9 + NH - 1}"
dash.conditional_formatting.add(
    data_range,
    CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
               fill=PatternFill("solid", fgColor=C_LIGHT_GREEN))
)
dash.conditional_formatting.add(
    data_range,
    CellIsRule(operator="between", formula=["0.5", "0.7999"],
               fill=PatternFill("solid", fgColor=C_LORANGE))
)
dash.conditional_formatting.add(
    data_range,
    CellIsRule(operator="lessThan", formula=["0.5"],
               fill=PatternFill("solid", fgColor=C_LRED))
)

# ── Per-habit yearly trend charts ────────────────────────────────
chart_start_row = overall_row + 3

# Line chart: monthly consistency per habit
chart = LineChart()
chart.title = "Monthly Consistency Trend"
chart.style = 10
chart.height = 15
chart.width = 28
chart.y_axis.title = "Consistency %"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1
chart.y_axis.numFmt = "0%"

cats = Reference(dash, min_col=3, max_col=14, min_row=8)
for h_i in range(NH):
    row = 9 + h_i
    data = Reference(dash, min_col=3, max_col=14, min_row=row, max_row=row)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.series[h_i].name = HABITS[h_i]["name"]

chart.set_categories(cats)
chart.legend.position = "b"
dash.add_chart(chart, f"A{chart_start_row}")

# Bar chart: per-habit year average
bar_start = chart_start_row + 18
bar = BarChart()
bar.type = "col"
bar.title = "Year Average by Habit"
bar.style = 10
bar.height = 12
bar.width = 16
bar.y_axis.numFmt = "0%"
bar.y_axis.scaling.min = 0
bar.y_axis.scaling.max = 1

bar_cats = Reference(dash, min_col=1, min_row=9, max_row=9 + NH - 1)
bar_data = Reference(dash, min_col=15, min_row=8, max_row=9 + NH - 1)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.legend = None
dash.add_chart(bar, f"A{bar_start}")

# ── Dashboard: How-to instructions ──────────────────────────────
instr_row = bar_start + 16
instructions = [
    "HOW TO USE THIS DASHBOARD",
    "",
    "1. Go to the current month's tab (e.g., March)",
    "2. Scroll down to the DAILY LOG section",
    "3. Enter 1 in a habit column when you complete it for the day",
    "4. Weekly and monthly percentages auto-update!",
    "",
    "TARGETS EXPLAINED",
    f"  · Gym, Speaking, Coding → 5 days per week (hit 5 = 100%)",
    f"  · Wake Early, Eat Clean, Calories → daily (7/7 = 100%)",
    "",
    "TIP: In Google Sheets on iPhone, you can select the habit",
    "input cells and use Insert > Checkbox for tap-to-toggle input.",
    "",
    "COLOR CODING",
    "  · Green  = 80%+ (crushing it)",
    "  · Orange = 50-79% (room to improve)",
    "  · Red    = below 50% (needs attention)",
]

for i, line in enumerate(instructions):
    r = instr_row + i
    font = F_SUBHDR if line.startswith("HOW") or line.startswith("TARGET") or line.startswith("COLOR") or line.startswith("TIP") else F_DATA_DIM
    cell(dash, r, 1, line, font=font, fill=None, align=LEFT,
         border=None, merge_end_col=15)

dash.freeze_panes = "A9"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SAVE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

wb.save(OUTPUT)
print(f"✅ Created {OUTPUT}")
print(f"   → {12} month sheets + Dashboard")
print(f"   → {NH} habits with smart weekly targets")
print(f"   → Formulas: weekly %, monthly %, yearly trend")
print(f"   → Conditional formatting: green/orange/red")
print(f"   → Charts: line trend + bar comparison")
print(f"\n📱 To use on iPhone:")
print(f"   1. Upload {OUTPUT} to Google Drive")
print(f"   2. Open with Google Sheets")
print(f"   3. Open in the Google Sheets app on your phone")
print(f"   4. Navigate to the current month tab")
print(f"   5. Enter 1 when you complete a habit!")
